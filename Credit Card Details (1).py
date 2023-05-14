#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
get_ipython().run_line_magic('autosave', '1')


# In[2]:


excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "Credit Cards"
print(excel.sheetnames)
sheet.append(['Card Name', 'Joining Fee', 'Annual Fee', 'Benefits of The Card'])


# # All the benefits for each credit card in a single cell of Excel sheet 

# In[23]:


import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl

excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "Credit Cards"
print(excel.sheetnames)
sheet.append(['Card Name', 'Joining Fee', 'Annual Fee', 'Benefits of The Card'])

source = requests.get('https://www.icicibank.com/card/credit-cards/credit-card')
source.raise_for_status()
soup = BeautifulSoup(source.text, 'html.parser')

cards = soup.find('div', class_='compare-account account-variants-single-card').find_all('div', class_='account-variants-card-with-img')

for card in cards:
    #Find the title element for the card
    card_name_elem = card.find('div', class_='titie-wrapper')
    #Find the joining fee for the card
    card_join_fee = card.find('span', class_="joining-fee").strong.text
    #finding benefits of card
    benefits_tags = card.find('ul').find_all('li')
    if card_name_elem is not None:
        #Find the card name if it is in an h3 tag
        card_name_h3 = card_name_elem.find('h3')
        if card_name_h3 is not None:
            card_name = card_name_h3.text.strip()
            print('Card Name:',card_name)
            print('Joining Fee:',card_join_fee)
            card_annual_fee = card.find('span',class_="line").strong
            if card_annual_fee is not None:
                print('Annual Fee:',card_annual_fee.text)
                print('\t')  
            else:
                continue
            benefits = "\n".join([tag.text for tag in benefits_tags])
            sheet.append([card_name,card_join_fee,card_annual_fee.text, benefits])
            print('Benefits of The Card:')
            print(benefits)
            print('\n\n')
        else:
            #Find the card name if it is in an h4 tag
            card_name_h4 = card_name_elem.find('h4')
            if card_name_h4 is not None:
                card_name = card_name_h4.text.strip()
                print('Card Name:',card_name)
                print('Joining Fee:',card_join_fee)
                card_annual_fee = card.find('span',class_="line").strong
                if card_annual_fee is not None:
                    print('Annual Fee:',card_annual_fee.text)
                    print('\t')  
                else:
                    continue
                benefits = "\n".join([tag.text for tag in benefits_tags])
                sheet.append([card_name,card_join_fee,card_annual_fee.text,benefits])
                print('Benefits of The Card:')
                print(benefits)
                print('\n\n')
                
excel.save('Credit_Card .xlsx')


# # All the benefits for each credit card in multiple cells of Excel sheet 

# In[22]:


import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "Credit Cards"
print(excel.sheetnames)
sheet.append(['Card Name', 'Joining Fee', 'Annual Fee', 'Benefits of The Card'])

import requests
from bs4 import BeautifulSoup

source = requests.get('https://www.icicibank.com/card/credit-cards/credit-card')
source.raise_for_status()
soup = BeautifulSoup(source.text, 'html.parser')

cards = soup.find('div', class_='compare-account account-variants-single-card').find_all('div', class_='account-variants-card-with-img')

for card in cards:
    #Find the name element for the card
    card_name_elem = card.find('div', class_='titie-wrapper')
    #Find the joining fee for the card
    card_join_fee = card.find('span', class_="joining-fee").strong.text
    #finding benefits of card
    benefits_tags = card.find('ul').find_all('li')
    if card_name_elem is not None:
        #Find the card name if it is in an h3 tag
        card_name_h3 = card_name_elem.find('h3')
        if card_name_h3 is not None:
            card_name = card_name_h3.text.strip()
            print('Card Name:',card_name)
            print('Joining Fee:',card_join_fee)
            card_annual_fee = card.find('span',class_="line").strong
            if card_annual_fee is not None:
                print('Annual Fee:',card_annual_fee.text)
                print('\t')  
            else:
                continue
            print('Benefits of The Card:')
            for tag in benefits_tags:
                print(tag.text)
                sheet.append([card_name,card_join_fee,card_annual_fee.text,tag.text])
            print('\n\n')
        else:
            #Find the card name if it is in an h4 tag
            card_name_h4 = card_name_elem.find('h4')
            if card_name_h4 is not None:
                card_name = card_name_h4.text.strip()
                print('Card Name:',card_name)
                print('Joining Fee:',card_join_fee)
                card_annual_fee = card.find('span',class_="line").strong
                if card_annual_fee is not None:
                    print('Annual Fee:',card_annual_fee.text)
                    print('\t')  
                else:
                    continue
                print('Benefits of The Card:')
                for tag in benefits_tags:
                    print(tag.text)
                    sheet.append([card_name,card_join_fee,card_annual_fee.text,tag.text])
                print('\n\n')
    
                
excel.save('Credit Card Details.xlsx')


# In[21]:


import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl

excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "Credit Cards"
sheet.append(['Card Name', 'Joining Fee', 'Annual Fee', 'Benefits of The Card'])

source = requests.get('https://www.icicibank.com/card/credit-cards/credit-card')
source.raise_for_status()
soup = BeautifulSoup(source.text, 'html.parser')

cards = soup.find('div', class_='compare-account account-variants-single-card').find_all('div', class_='account-variants-card-with-img')

for card in cards:
    #Find the title element for the card
    card_name_elem = card.find('div', class_='titie-wrapper')
    #Find the joining fee for the card
    card_join_fee = card.find('span', class_="joining-fee").strong.text
    #finding benefits of card
    benefits_tags = card.find('ul').find_all('li')
    if card_name_elem is not None:
        #Find the card name if it is in an h3 tag
        card_name_h3 = card_name_elem.find('h3')
        if card_name_h3 is not None:
            card_name = card_name_h3.text.strip()
            print('Card Name:',card_name)
            print('Joining Fee:',card_join_fee)
            card_annual_fee = card.find('span',class_="line").strong
            if card_annual_fee is not None:
                print('Annual Fee:',card_annual_fee.text)
                print('\t')  
            else:
                continue
            # Join the benefits tags using ';' as a delimiter
            benefits = ';'.join([tag.text for tag in benefits_tags])
            sheet.append([card_name, card_join_fee, card_annual_fee.text, benefits])
            print('Benefits of The Card:', benefits)
            print('\n\n')
        else:
            #Find the card name if it is in an h4 tag
            card_name_h4 = card_name_elem.find('h4')
            if card_name_h4 is not None:
                card_name = card_name_h4.text.strip()
                print('Card Name:',card_name)
                print('Joining Fee:',card_join_fee)
                card_annual_fee = card.find('span',class_="line").strong
                if card_annual_fee is not None:
                    print('Annual Fee:',card_annual_fee.text)
                    print('\t')  
                else:
                    continue
                # Join the benefits tags using ';' as a delimiter
                benefits = ';'.join([tag.text for tag in benefits_tags])
                sheet.append([card_name, card_join_fee, card_annual_fee.text, benefits])
                print('Benefits of The Card:', benefits)
                print('\n\n')
                
excel.save('Credit.xlsx')


# In[ ]:




